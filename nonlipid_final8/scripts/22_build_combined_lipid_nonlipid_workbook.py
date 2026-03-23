from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion")
LIPID_XLSX = ROOT / "supplement_lipid_final8" / "lipid_final8_supplement_workbook.xlsx"
NONLIPID_XLSX = ROOT / "supplement_nonlipid_final8" / "nonlipid_final8_supplement_workbook.xlsx"

OUT_DIR = ROOT / "supplement_combined_lipid_nonlipid_final8"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_XLSX = OUT_DIR / "metabolic_lipid_nonlipid_final8_combined_supplement_workbook.xlsx"
OUT_TXT = OUT_DIR / "metabolic_lipid_nonlipid_final8_combined_overview.txt"


def read_sheet(path, sheet_name, index_col=None):
    return pd.read_excel(path, sheet_name=sheet_name, index_col=index_col)


def add_module(df, module, stage_label=None):
    out = df.copy()
    out.insert(0, "module", module)
    if stage_label is not None:
        out.insert(1, "stage_label", stage_label)
    return out


def matrix_to_long(df, module, stage_label, value_name):
    out = df.copy()
    out.index.name = "row_trait"
    out.columns.name = "col_trait"
    out = out.stack(future_stack=True).reset_index(name=value_name)
    out.insert(0, "module", module)
    out.insert(1, "stage_label", stage_label)
    return out


def factor_rg_to_long(df, module):
    out = df.copy()
    out.index.name = "row_trait"
    out.columns.name = "col_trait"
    out = out.stack(future_stack=True).reset_index(name="rg")
    out.insert(0, "module", module)
    return out


def align_columns(frames):
    all_cols = []
    for df in frames:
        for col in df.columns:
            if col not in all_cols:
                all_cols.append(col)
    return [df.reindex(columns=all_cols) for df in frames]


shared_main_qc = read_sheet(LIPID_XLSX, "S2_Main112_QC")
shared_main_s = read_sheet(LIPID_XLSX, "S3_Main112_S", index_col=0)
shared_main_rg = read_sheet(LIPID_XLSX, "S4_Main112_rg", index_col=0)

lipid_stage = read_sheet(LIPID_XLSX, "S1_StageCriteria")
nonlipid_stage = read_sheet(NONLIPID_XLSX, "01_StageCriteria")
combined_stage = pd.concat(
    align_columns([
        add_module(lipid_stage, "lipid"),
        add_module(nonlipid_stage, "nonlipid"),
    ]),
    ignore_index=True,
)

lipid_lineage = read_sheet(LIPID_XLSX, "S5_Lineage_112to8")
nonlipid_lineage = read_sheet(NONLIPID_XLSX, "05_Lineage_112to8")
combined_lineage = pd.concat(
    align_columns([
        add_module(lipid_lineage, "lipid"),
        add_module(nonlipid_lineage, "nonlipid"),
    ]),
    ignore_index=True,
)

lipid_candidates = read_sheet(LIPID_XLSX, "S6_Lipid94_Candidates")
nonlipid_candidates = read_sheet(NONLIPID_XLSX, "07_Nonlipid18_Candidates")
combined_candidates = pd.concat(
    align_columns([
        add_module(lipid_candidates, "lipid", "candidate_universe"),
        add_module(nonlipid_candidates, "nonlipid", "candidate_universe"),
    ]),
    ignore_index=True,
)

lipid_compact_review = read_sheet(LIPID_XLSX, "S7_Compact20_Review")
nonlipid_compact_review = read_sheet(NONLIPID_XLSX, "09_Compact15_Review")
combined_compact_review = pd.concat(
    align_columns([
        add_module(lipid_compact_review, "lipid", "compact_review"),
        add_module(nonlipid_compact_review, "nonlipid", "compact_review"),
    ]),
    ignore_index=True,
)

lipid_compact_ldsc = read_sheet(LIPID_XLSX, "S8_Compact20_LDSC")
nonlipid_compact_ldsc = read_sheet(NONLIPID_XLSX, "10_Compact15_LDSC")
combined_compact_ldsc = pd.concat(
    align_columns([
        add_module(lipid_compact_ldsc, "lipid", "compact_ldsc"),
        add_module(nonlipid_compact_ldsc, "nonlipid", "compact_ldsc"),
    ]),
    ignore_index=True,
)

lipid_compact_s = read_sheet(LIPID_XLSX, "S9_Compact20_S", index_col=0)
nonlipid_compact_s = read_sheet(NONLIPID_XLSX, "11_Compact15_S", index_col=0)
combined_compact_s = pd.concat(
    [
        matrix_to_long(lipid_compact_s, "lipid", "compact_panel", "s_estimate"),
        matrix_to_long(nonlipid_compact_s, "nonlipid", "compact_panel", "s_estimate"),
    ],
    ignore_index=True,
)

lipid_compact_rg = read_sheet(LIPID_XLSX, "S10_Compact20_rg", index_col=0)
nonlipid_compact_rg = read_sheet(NONLIPID_XLSX, "12_Compact15_rg", index_col=0)
combined_compact_rg = pd.concat(
    [
        matrix_to_long(lipid_compact_rg, "lipid", "compact_panel", "rg"),
        matrix_to_long(nonlipid_compact_rg, "nonlipid", "compact_panel", "rg"),
    ],
    ignore_index=True,
)

lipid_compact_efa = read_sheet(LIPID_XLSX, "S19_Compact20_EFA_Fit")
nonlipid_compact_efa = read_sheet(NONLIPID_XLSX, "13_Compact15_EFA_Fit")
combined_compact_efa = pd.concat(
    align_columns([
        add_module(lipid_compact_efa, "lipid", "compact_efa_fit"),
        add_module(nonlipid_compact_efa, "nonlipid", "compact_efa_fit"),
    ]),
    ignore_index=True,
)

lipid_compact_esem = read_sheet(LIPID_XLSX, "S20_Compact20_ESEM_Fit")
nonlipid_compact_esem = read_sheet(NONLIPID_XLSX, "14_Compact15_ESEM_Fit")
combined_compact_esem = pd.concat(
    align_columns([
        add_module(lipid_compact_esem, "lipid", "compact_esem_fit"),
        add_module(nonlipid_compact_esem, "nonlipid", "compact_esem_fit"),
    ]),
    ignore_index=True,
)

lipid_ultra_review = read_sheet(LIPID_XLSX, "S11_UltraPure10_Review")
lipid_final_review = read_sheet(LIPID_XLSX, "S12_Final8_Review")
nonlipid_ultra_review = read_sheet(NONLIPID_XLSX, "15_Ultra8_Review")
combined_refinement_review = pd.concat(
    align_columns([
        add_module(lipid_ultra_review, "lipid", "ultra10_review"),
        add_module(lipid_final_review, "lipid", "final8_review"),
        add_module(nonlipid_ultra_review, "nonlipid", "ultra8_review"),
    ]),
    ignore_index=True,
)

lipid_final_manifest = read_sheet(LIPID_XLSX, "S13_Final8_Manifest")
nonlipid_final_manifest = read_sheet(NONLIPID_XLSX, "16_Ultra8_Manifest")
combined_final_manifest = pd.concat(
    align_columns([
        add_module(lipid_final_manifest, "lipid", "final_manifest"),
        add_module(nonlipid_final_manifest, "nonlipid", "final_manifest"),
    ]),
    ignore_index=True,
)

lipid_final_ldsc = read_sheet(LIPID_XLSX, "S14_Final8_LDSC")
nonlipid_final_ldsc = read_sheet(NONLIPID_XLSX, "17_Ultra8_LDSC")
combined_final_ldsc = pd.concat(
    align_columns([
        add_module(lipid_final_ldsc, "lipid", "final_ldsc"),
        add_module(nonlipid_final_ldsc, "nonlipid", "final_ldsc"),
    ]),
    ignore_index=True,
)

lipid_final_s = read_sheet(LIPID_XLSX, "S15_Final8_S", index_col=0)
nonlipid_final_s = read_sheet(NONLIPID_XLSX, "18_Ultra8_S", index_col=0)
combined_final_s = pd.concat(
    [
        matrix_to_long(lipid_final_s, "lipid", "final_panel", "s_estimate"),
        matrix_to_long(nonlipid_final_s, "nonlipid", "final_panel", "s_estimate"),
    ],
    ignore_index=True,
)

lipid_final_rg = read_sheet(LIPID_XLSX, "S16_Final8_rg", index_col=0)
nonlipid_final_rg = read_sheet(NONLIPID_XLSX, "19_Ultra8_rg", index_col=0)
combined_final_rg = pd.concat(
    [
        matrix_to_long(lipid_final_rg, "lipid", "final_panel", "rg"),
        matrix_to_long(nonlipid_final_rg, "nonlipid", "final_panel", "rg"),
    ],
    ignore_index=True,
)

lipid_final_efa = read_sheet(LIPID_XLSX, "S23_Final8_EFA_Fit")
nonlipid_final_efa = read_sheet(NONLIPID_XLSX, "20_Ultra8_EFA_Fit")
combined_final_efa = pd.concat(
    align_columns([
        add_module(lipid_final_efa, "lipid", "final_efa_fit"),
        add_module(nonlipid_final_efa, "nonlipid", "final_efa_fit"),
    ]),
    ignore_index=True,
)

lipid_final_esem = read_sheet(LIPID_XLSX, "S17_Final8_ESEM_Fit")
nonlipid_final_esem = read_sheet(NONLIPID_XLSX, "21_Ultra8_ESEM_Fit")
combined_final_esem = pd.concat(
    align_columns([
        add_module(lipid_final_esem, "lipid", "final_esem_fit"),
        add_module(nonlipid_final_esem, "nonlipid", "final_esem_fit"),
    ]),
    ignore_index=True,
)

lipid_loadings = read_sheet(LIPID_XLSX, "S18_Final8_Loadings")
nonlipid_loadings = read_sheet(NONLIPID_XLSX, "22_Ultra8_Loadings")
combined_loadings = pd.concat(
    align_columns([
        add_module(lipid_loadings, "lipid", "final_loadings"),
        add_module(nonlipid_loadings, "nonlipid", "final_loadings"),
    ]),
    ignore_index=True,
)

lipid_fgwas = read_sheet(LIPID_XLSX, "S24_FactorGWAS_Files")
nonlipid_fgwas = read_sheet(NONLIPID_XLSX, "23_FactorGWAS_Files")
combined_fgwas = pd.concat(
    align_columns([
        add_module(lipid_fgwas, "lipid", "factor_gwas_files"),
        add_module(nonlipid_fgwas, "nonlipid", "factor_gwas_files"),
    ]),
    ignore_index=True,
)

lipid_factor_uni = read_sheet(LIPID_XLSX, "S25_FactorLDSC_Uni")
nonlipid_factor_uni = read_sheet(NONLIPID_XLSX, "24_FactorLDSC_Uni")
combined_factor_uni = pd.concat(
    align_columns([
        add_module(lipid_factor_uni, "lipid", "factor_ldsc_uni"),
        add_module(nonlipid_factor_uni, "nonlipid", "factor_ldsc_uni"),
    ]),
    ignore_index=True,
)

lipid_factor_bi = read_sheet(LIPID_XLSX, "S26_FactorLDSC_Bi")
nonlipid_factor_bi = read_sheet(NONLIPID_XLSX, "25_FactorLDSC_Bi")
combined_factor_bi = pd.concat(
    align_columns([
        add_module(lipid_factor_bi, "lipid", "factor_ldsc_bi"),
        add_module(nonlipid_factor_bi, "nonlipid", "factor_ldsc_bi"),
    ]),
    ignore_index=True,
)

lipid_factor_rg = read_sheet(LIPID_XLSX, "S27_FactorLDSC_rg", index_col=0)
nonlipid_factor_rg = read_sheet(NONLIPID_XLSX, "26_FactorLDSC_rg", index_col=0)
combined_factor_rg = pd.concat(
    [
        factor_rg_to_long(lipid_factor_rg, "lipid"),
        factor_rg_to_long(nonlipid_factor_rg, "nonlipid"),
    ],
    ignore_index=True,
)

nonlipid_group_include = read_sheet(NONLIPID_XLSX, "06_GroupInclusion")
nonlipid_grouping = read_sheet(NONLIPID_XLSX, "08_Grouping")

index_df = pd.DataFrame(
    [
        ("00_Index", "Combined workbook directory and sheet descriptions."),
        ("01_StageCriteria", "Lipid and nonlipid stage-selection criteria merged into one sheet with a module column."),
        ("02_Main112_QC_Shared", "Shared 112-trait QC table used by both modules."),
        ("03_Main112_S_Shared", "Shared 112-trait multivariate LDSC S matrix."),
        ("04_Main112_rg_Shared", "Shared 112-trait multivariate LDSC genetic-correlation matrix."),
        ("05_Lineage_Combined", "Trait-level lineage for both modules merged with aligned columns."),
        ("06_CandidateUniverse_Combined", "Lipid and nonlipid candidate-universe tables merged with a module column."),
        ("07_CompactReview_Combined", "Compact-stage review tables merged with a module column."),
        ("08_CompactLDSC_Combined", "Compact-stage multivariate LDSC summaries merged with a module column."),
        ("09_CompactS_Long", "Compact-stage S matrices converted to long format and merged across modules."),
        ("10_Compactrg_Long", "Compact-stage rg matrices converted to long format and merged across modules."),
        ("11_CompactEFA_Fit", "Compact-stage EFA fit summaries merged across modules."),
        ("12_CompactESEM_Fit", "Compact-stage ESEM fit summaries merged across modules."),
        ("13_RefinementReview_Combined", "Later-stage review tables combined, including lipid ultra10/final8 and nonlipid ultra8 reviews."),
        ("14_FinalManifest_Combined", "Final 8-trait manifests for lipid and nonlipid merged into one table."),
        ("15_FinalLDSC_Combined", "Final panel multivariate LDSC summaries merged across modules."),
        ("16_FinalS_Long", "Final-panel S matrices converted to long format and merged across modules."),
        ("17_Finalrg_Long", "Final-panel rg matrices converted to long format and merged across modules."),
        ("18_FinalEFA_Fit", "Final-panel EFA fit summaries merged across modules."),
        ("19_FinalESEM_Fit", "Final-panel ESEM fit summaries merged across modules."),
        ("20_FinalLoadings_Combined", "Final 3-factor loading tables merged across modules."),
        ("21_FactorGWAS_Files", "Final factor-GWAS merged-file manifests merged across modules."),
        ("22_FactorLDSC_Uni", "Final factor-level univariate LDSC details merged across modules."),
        ("23_FactorLDSC_Bi", "Final factor-level bivariate LDSC details merged across modules."),
        ("24_FactorLDSC_rg_Long", "Final factor-level rg matrices converted to long format and merged across modules."),
        ("25_Nonlipid_GroupInclusion", "Nonlipid-specific group-inclusion table kept separately for interpretation."),
        ("26_Nonlipid_Grouping", "Nonlipid-specific grouping table kept separately for interpretation."),
    ],
    columns=["sheet_name", "description"],
)

with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    index_df.to_excel(writer, sheet_name="00_Index", index=False)
    combined_stage.to_excel(writer, sheet_name="01_StageCriteria", index=False)
    shared_main_qc.to_excel(writer, sheet_name="02_Main112_QC_Shared", index=False)
    shared_main_s.to_excel(writer, sheet_name="03_Main112_S_Shared")
    shared_main_rg.to_excel(writer, sheet_name="04_Main112_rg_Shared")
    combined_lineage.to_excel(writer, sheet_name="05_Lineage_Combined", index=False)
    combined_candidates.to_excel(writer, sheet_name="06_CandidateUniverse_Combined", index=False)
    combined_compact_review.to_excel(writer, sheet_name="07_CompactReview_Combined", index=False)
    combined_compact_ldsc.to_excel(writer, sheet_name="08_CompactLDSC_Combined", index=False)
    combined_compact_s.to_excel(writer, sheet_name="09_CompactS_Long", index=False)
    combined_compact_rg.to_excel(writer, sheet_name="10_Compactrg_Long", index=False)
    combined_compact_efa.to_excel(writer, sheet_name="11_CompactEFA_Fit", index=False)
    combined_compact_esem.to_excel(writer, sheet_name="12_CompactESEM_Fit", index=False)
    combined_refinement_review.to_excel(writer, sheet_name="13_RefinementReview_Combined", index=False)
    combined_final_manifest.to_excel(writer, sheet_name="14_FinalManifest_Combined", index=False)
    combined_final_ldsc.to_excel(writer, sheet_name="15_FinalLDSC_Combined", index=False)
    combined_final_s.to_excel(writer, sheet_name="16_FinalS_Long", index=False)
    combined_final_rg.to_excel(writer, sheet_name="17_Finalrg_Long", index=False)
    combined_final_efa.to_excel(writer, sheet_name="18_FinalEFA_Fit", index=False)
    combined_final_esem.to_excel(writer, sheet_name="19_FinalESEM_Fit", index=False)
    combined_loadings.to_excel(writer, sheet_name="20_FinalLoadings_Combined", index=False)
    combined_fgwas.to_excel(writer, sheet_name="21_FactorGWAS_Files", index=False)
    combined_factor_uni.to_excel(writer, sheet_name="22_FactorLDSC_Uni", index=False)
    combined_factor_bi.to_excel(writer, sheet_name="23_FactorLDSC_Bi", index=False)
    combined_factor_rg.to_excel(writer, sheet_name="24_FactorLDSC_rg_Long", index=False)
    nonlipid_group_include.to_excel(writer, sheet_name="25_Nonlipid_GroupInclusion", index=False)
    nonlipid_grouping.to_excel(writer, sheet_name="26_Nonlipid_Grouping", index=False)

wb = load_workbook(OUT_XLSX)
header_fill = PatternFill("solid", fgColor="D9EAF7")
header_font = Font(bold=True)
top_align = Alignment(vertical="top", wrap_text=True)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = top_align
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells[:200]:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 42)

ws = wb["00_Index"]
ws["D1"] = "notes"
ws["D2"] = "This combined workbook preserves the original lipid and nonlipid workbooks and merges same-analysis outputs into shared sheets using a module column."
ws["D3"] = "Shared Main112 QC/S/rg sheets are included once because both modules use the same Z > 4 starting universe."
ws["D4"] = "Matrix-style module-specific outputs are converted to long format so lipid and nonlipid results can be reviewed in the same sheet."
ws["D5"] = "Nonlipid-specific GroupInclusion and Grouping sheets are retained separately because there is no direct lipid analogue."
ws["D6"] = "Lipid FactorGWAS metadata does not include q_snp_status; nonlipid metadata records q_snp_status = not_generated."
for cell_ref in ["D2", "D3", "D4", "D5", "D6"]:
    ws[cell_ref].alignment = top_align
ws.column_dimensions["D"].width = 95

wb.save(OUT_XLSX)

OUT_TXT.write_text(
    "\n".join(
        [
            f"combined_workbook\t{OUT_XLSX}",
            f"source_lipid_workbook\t{LIPID_XLSX}",
            f"source_nonlipid_workbook\t{NONLIPID_XLSX}",
            "merge_rule\tsame-analysis sheets merged with a module column",
            "matrix_rule\tmodule-specific matrices converted to long format before merge",
            "shared_main112_rule\tshared QC/S/rg retained once because the source universe is identical",
        ]
    )
    + "\n",
    encoding="utf-8",
)

print(f"Wrote workbook: {OUT_XLSX}")
print(f"Wrote overview: {OUT_TXT}")
