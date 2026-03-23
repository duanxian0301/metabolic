from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion")
OUT_DIR = ROOT / "supplement_nonlipid_final8"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_XLSX = OUT_DIR / "nonlipid_final8_supplement_workbook.xlsx"


def read_tsv(path):
    return pd.read_csv(path, sep="\t")


def read_csv_index(path):
    return pd.read_csv(path, index_col=0)


main_manifest = read_tsv(ROOT / "step1_ldsc_results" / "trait_manifest.tsv")
main_qc = read_tsv(ROOT / "step1_ldsc_results" / "efa_selection" / "trait_qc_and_selection.tsv")
main_s = read_csv_index(ROOT / "step1_ldsc_results" / "Main_Zgt4_nonproportion_S_matrix.csv")
main_rg = read_csv_index(ROOT / "step1_ldsc_results" / "Main_Zgt4_nonproportion_rg_matrix.csv")

nonlipid18 = read_tsv(ROOT / "nonlipid_module_from_full_manifest" / "nonlipid_module_candidates.tsv")
group_include = read_tsv(ROOT / "nonlipid_module_from_full_manifest" / "nonlipid_group_inclusion_table.tsv")
grouping = read_tsv(ROOT / "nonlipid_module_from_full_manifest" / "nonlipid_module_grouping.tsv")
compact15_review = read_tsv(ROOT / "nonlipid_module_from_full_manifest" / "compact_panel_review" / "nonlipid_module_compact_review.tsv")
ultra8_review = read_tsv(ROOT / "nonlipid_module_from_full_manifest" / "compact_panel_review" / "ultra_pure_3factor_review" / "nonlipid_module_ultrapure3_review.tsv")
final8_manifest = read_tsv(ROOT / "final_model_nonlipid_module_final8" / "final8_trait_manifest.tsv")

compact15_summary = read_tsv(ROOT / "step16_ldsc_nonlipid_module_compact15" / "nonlipid_module_compact15_ldsc_summary.tsv")
compact15_s = read_csv_index(ROOT / "step16_ldsc_nonlipid_module_compact15" / "nonlipid_module_compact15_S_matrix.csv")
compact15_rg = read_csv_index(ROOT / "step16_ldsc_nonlipid_module_compact15" / "nonlipid_module_compact15_rg_matrix.csv")
compact15_efa_fit = read_tsv(ROOT / "step17_efa_esem_nonlipid_module_compact15" / "EFA_minres_fit_summary.tsv")
compact15_esem_fit = read_tsv(ROOT / "step17_efa_esem_nonlipid_module_compact15" / "nonlipid_module_compact15_esem_target_summary.tsv")

ultra8_summary = read_tsv(ROOT / "step19_ldsc_nonlipid_module_ultrapure8" / "nonlipid_module_ultrapure8_ldsc_summary.tsv")
ultra8_s = read_csv_index(ROOT / "step19_ldsc_nonlipid_module_ultrapure8" / "nonlipid_module_ultrapure8_S_matrix.csv")
ultra8_rg = read_csv_index(ROOT / "step19_ldsc_nonlipid_module_ultrapure8" / "nonlipid_module_ultrapure8_rg_matrix.csv")
ultra8_efa_fit = read_tsv(ROOT / "step20_efa_esem_nonlipid_module_ultrapure8" / "EFA_minres_fit_summary.tsv")
ultra8_esem_fit = read_tsv(ROOT / "step20_efa_esem_nonlipid_module_ultrapure8" / "nonlipid_module_ultrapure8_esem_summary.tsv")
ultra8_loadings = read_tsv(ROOT / "step20_efa_esem_nonlipid_module_ultrapure8" / "ALL_3factor_loadings.tsv")

fgwas_manifest = read_tsv(ROOT / "step22_native_wsl_usergwas_nonlipid8_results" / "merged_nonlipid_final8" / "standard_txt" / "nonlipid_final8_standard_txt_metadata.tsv")
factor_ldsc_uni = read_tsv(ROOT / "supplement_nonlipid_final8" / "Supplementary_Table_S5_nonlipid_final8_univariate_ldsc_detailed.tsv")
factor_ldsc_bi = read_tsv(ROOT / "supplement_nonlipid_final8" / "Supplementary_Table_S6_nonlipid_final8_bivariate_ldsc_detailed.tsv")
factor_ldsc_rg = pd.read_csv(ROOT / "step23_ldsc_validation_nonlipid_final8" / "nonlipid_final8_factor_rg_matrix.tsv", sep="\t")
factor_ldsc_rg.index = ["nonlipid8_F1", "nonlipid8_F2", "nonlipid8_F3"]


lineage = main_qc.merge(main_manifest, on=["study_accession", "trait_code", "biomarker_name", "group"], how="left")
nonlipid_set = set(nonlipid18["trait_code"])
compact_cols = ["trait_code", "selection_status", "selection_reason", "proposed_module", "marker_role", "high_rg_links_gt_0p95"]
ultra_cols = ["trait_code", "selection_status", "selection_reason", "ultra_pure_factor", "same_primary", "mean_top_abs", "mean_gap"]

lineage["stage_18_status"] = lineage["trait_code"].map(lambda x: "kept" if x in nonlipid_set else "dropped")
lineage["stage_18_reason"] = lineage["group"].map(
    lambda g: "Retained in nonlipid candidate universe." if g in set(nonlipid18["group"]) else "Dropped because trait belongs to the excluded lipid/lipoprotein universe."
)

lineage = lineage.merge(
    compact15_review[compact_cols].rename(columns={
        "selection_status": "stage_15_status",
        "selection_reason": "stage_15_reason",
        "proposed_module": "stage_15_module",
        "marker_role": "stage_15_marker_role",
        "high_rg_links_gt_0p95": "stage_15_high_rg_links_gt_0p95",
    }),
    on="trait_code", how="left"
)
lineage = lineage.merge(
    ultra8_review[ultra_cols].rename(columns={
        "selection_status": "stage_8_status",
        "selection_reason": "stage_8_reason",
        "ultra_pure_factor": "stage_8_factor",
        "same_primary": "stage_8_same_primary",
        "mean_top_abs": "stage_8_mean_top_abs",
        "mean_gap": "stage_8_mean_gap",
    }),
    on="trait_code", how="left"
)

for col in ["stage_15_status", "stage_8_status"]:
    lineage[col] = lineage[col].fillna("not_entered")

lineage["final_factor_name"] = lineage["trait_code"].map(dict(zip(final8_manifest["trait_code"], final8_manifest["final_factor_name"])))

criteria = pd.DataFrame([
    {
        "stage": "Stage 1",
        "input_n": "all Main_Zgt4_nonproportion traits",
        "output_n": 112,
        "selection_standard": "Retain traits with single-trait h2 Z > 4 in the original QC table.",
        "dropped_n": "not shown here",
        "notes": "This is the shared main-analysis universe."
    },
    {
        "stage": "Stage 2",
        "input_n": 112,
        "output_n": 18,
        "selection_standard": "Restrict to nonlipid groups and exclude lipid/lipoprotein groups.",
        "dropped_n": 94,
        "notes": "Included amino acids, glycolysis-related metabolites, ketone bodies, fluid balance, and inflammation."
    },
    {
        "stage": "Stage 3",
        "input_n": 18,
        "output_n": 15,
        "selection_standard": "Compact panel preserving amino-acid, ketone, glycolysis, and bridge traits while removing obvious redundancy and low-priority support markers.",
        "dropped_n": 3,
        "notes": "Ala, Ile, and Acetone were dropped at compact review."
    },
    {
        "stage": "Stage 4",
        "input_n": 15,
        "output_n": 8,
        "selection_standard": "Ultrapure 3-factor panel chosen by stable primary loading, loading-gap support, and compact-subset ESEM search favoring CFI >= 0.95 and SRMR < 0.08.",
        "dropped_n": 7,
        "notes": "This produced the final nonlipid 8-trait 3-factor panel."
    },
    {
        "stage": "Stage 5",
        "input_n": 8,
        "output_n": "3 factors",
        "selection_standard": "Final downstream model: 3-factor target-rotation ESEM followed by native WSL userGWAS and factor-level LDSC validation.",
        "dropped_n": 0,
        "notes": "Q_SNP chunk outputs were not generated, but F1/F2/F3 userGWAS outputs were complete."
    },
])

index_df = pd.DataFrame([
    ("00_Index", "Workbook directory and sheet descriptions."),
    ("01_StageCriteria", "Nonlipid final8 selection stages from 112 traits to final 8 traits."),
    ("02_Main112_QC", "Initial 112-trait QC table from the main Z>4 nonproportion universe."),
    ("03_Main112_S", "112-trait multivariate LDSC S matrix."),
    ("04_Main112_rg", "112-trait multivariate LDSC genetic correlation matrix."),
    ("05_Lineage_112to8", "Trait-by-trait lineage across the nonlipid 18, compact 15, and final 8 stages."),
    ("06_GroupInclusion", "Group-level inclusion table defining the nonlipid candidate universe."),
    ("07_Nonlipid18_Candidates", "Nonlipid candidate universe (18 traits)."),
    ("08_Grouping", "Module and subgroup organization for nonlipid candidates."),
    ("09_Compact15_Review", "15-trait compact nonlipid review with keep/drop decisions and reasons."),
    ("10_Compact15_LDSC", "15-trait compact panel multivariate LDSC summary."),
    ("11_Compact15_S", "15-trait compact panel S matrix."),
    ("12_Compact15_rg", "15-trait compact panel rg matrix."),
    ("13_Compact15_EFA_Fit", "Compact 15-trait EFA fit summary."),
    ("14_Compact15_ESEM_Fit", "Compact 15-trait ESEM model comparison summary."),
    ("15_Ultra8_Review", "Final nonlipid 8-trait ultrapure review with keep/drop decisions."),
    ("16_Ultra8_Manifest", "Final 8-trait manifest and factor assignment."),
    ("17_Ultra8_LDSC", "Final 8-trait multivariate LDSC summary."),
    ("18_Ultra8_S", "Final 8-trait S matrix."),
    ("19_Ultra8_rg", "Final 8-trait rg matrix."),
    ("20_Ultra8_EFA_Fit", "Final 8-trait EFA fit summary."),
    ("21_Ultra8_ESEM_Fit", "Final 8-trait ESEM fit summary."),
    ("22_Ultra8_Loadings", "ALL dataset 3-factor loadings for the final 8-trait model."),
    ("23_FactorGWAS_Files", "Final merged userGWAS outputs and standard TXT exports."),
    ("24_FactorLDSC_Uni", "Final factor-level univariate LDSC summary."),
    ("25_FactorLDSC_Bi", "Final factor-level bivariate LDSC summary."),
    ("26_FactorLDSC_rg", "Final factor-level rg matrix.")
], columns=["sheet_name", "description"])

with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    index_df.to_excel(writer, sheet_name="00_Index", index=False)
    criteria.to_excel(writer, sheet_name="01_StageCriteria", index=False)
    main_qc.to_excel(writer, sheet_name="02_Main112_QC", index=False)
    main_s.to_excel(writer, sheet_name="03_Main112_S")
    main_rg.to_excel(writer, sheet_name="04_Main112_rg")
    lineage.to_excel(writer, sheet_name="05_Lineage_112to8", index=False)
    group_include.to_excel(writer, sheet_name="06_GroupInclusion", index=False)
    nonlipid18.to_excel(writer, sheet_name="07_Nonlipid18_Candidates", index=False)
    grouping.to_excel(writer, sheet_name="08_Grouping", index=False)
    compact15_review.to_excel(writer, sheet_name="09_Compact15_Review", index=False)
    compact15_summary.to_excel(writer, sheet_name="10_Compact15_LDSC", index=False)
    compact15_s.to_excel(writer, sheet_name="11_Compact15_S")
    compact15_rg.to_excel(writer, sheet_name="12_Compact15_rg")
    compact15_efa_fit.to_excel(writer, sheet_name="13_Compact15_EFA_Fit", index=False)
    compact15_esem_fit.to_excel(writer, sheet_name="14_Compact15_ESEM_Fit", index=False)
    ultra8_review.to_excel(writer, sheet_name="15_Ultra8_Review", index=False)
    final8_manifest.to_excel(writer, sheet_name="16_Ultra8_Manifest", index=False)
    ultra8_summary.to_excel(writer, sheet_name="17_Ultra8_LDSC", index=False)
    ultra8_s.to_excel(writer, sheet_name="18_Ultra8_S")
    ultra8_rg.to_excel(writer, sheet_name="19_Ultra8_rg")
    ultra8_efa_fit.to_excel(writer, sheet_name="20_Ultra8_EFA_Fit", index=False)
    ultra8_esem_fit.to_excel(writer, sheet_name="21_Ultra8_ESEM_Fit", index=False)
    ultra8_loadings.to_excel(writer, sheet_name="22_Ultra8_Loadings", index=False)
    fgwas_manifest.to_excel(writer, sheet_name="23_FactorGWAS_Files", index=False)
    factor_ldsc_uni.to_excel(writer, sheet_name="24_FactorLDSC_Uni", index=False)
    factor_ldsc_bi.to_excel(writer, sheet_name="25_FactorLDSC_Bi", index=False)
    factor_ldsc_rg.to_excel(writer, sheet_name="26_FactorLDSC_rg")

wb = load_workbook(OUT_XLSX)
header_fill = PatternFill("solid", fgColor="D9EAF7")
header_font = Font(bold=True)
top_align = Alignment(vertical="top", wrap_text=True)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = top_align
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells[:200]:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

if "00_Index" in wb.sheetnames:
    ws = wb["00_Index"]
    ws["D1"] = "notes"
    ws["D2"] = "Workbook includes the successful clean nonlipid route from 112 traits to the final 8-trait 3-factor model."
    ws["D3"] = "The 112-trait V matrix is omitted because it is not practical for Excel review; S and rg matrices are included instead."
    ws["D4"] = "Q_SNP chunk outputs were not generated in the native WSL userGWAS run; factor GWAS outputs for F1/F2/F3 were complete."
    ws["D2"].alignment = top_align
    ws["D3"].alignment = top_align
    ws["D4"].alignment = top_align
    ws.column_dimensions["D"].width = 90

wb.save(OUT_XLSX)
print(f"Wrote workbook: {OUT_XLSX}")
