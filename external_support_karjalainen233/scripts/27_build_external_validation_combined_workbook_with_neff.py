import csv
import shutil
from pathlib import Path

import openpyxl


ROOT = Path(r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion")
SOURCE_WORKBOOK = ROOT / "supplement_combined_lipid_nonlipid_final8" / "metabolic_lipid_nonlipid_final8_combined_supplement_workbook_with_external_validation.xlsx"
TARGET_WORKBOOK = ROOT / "supplement_combined_lipid_nonlipid_final8" / "metabolic_lipid_nonlipid_final8_combined_supplement_workbook_with_external_validation_and_neff.xlsx"

LIPID_META = ROOT / "step14_native_wsl_usergwas_final8_results" / "merged_lipid_final8" / "standard_txt" / "lipid_final8_standard_txt_metadata.tsv"
NONLIPID_META = ROOT / "step22_native_wsl_usergwas_nonlipid8_results" / "merged_nonlipid_final8" / "standard_txt" / "nonlipid_final8_standard_txt_metadata.tsv"
NEFF_SUMMARY = ROOT / "factor_standard_txt_neff_summary.tsv"


def read_tsv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def main():
    shutil.copyfile(SOURCE_WORKBOOK, TARGET_WORKBOOK)
    wb = openpyxl.load_workbook(TARGET_WORKBOOK)

    factor_meta = {}
    for module, path in [("lipid", LIPID_META), ("nonlipid", NONLIPID_META)]:
        for row in read_tsv(path):
            factor_meta[(module, row["factor"])] = row

    ws = wb["21_FactorGWAS_Files"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_index = {name: idx + 1 for idx, name in enumerate(headers)}

    for row_idx in range(2, ws.max_row + 1):
        module = ws.cell(row_idx, col_index["module"]).value
        factor = ws.cell(row_idx, col_index["factor"]).value
        meta = factor_meta[(module, factor)]
        ws.cell(row_idx, col_index["n_used"]).value = int(meta["n_used"])
        ws.cell(row_idx, col_index["notes"]).value = meta["notes"]
        if "q_snp_status" in col_index and "q_snp_status" in meta:
            ws.cell(row_idx, col_index["q_snp_status"]).value = meta["q_snp_status"]

    neff_rows = read_tsv(NEFF_SUMMARY)
    neff_sheet = "30_FactorGWAS_Neff"
    if neff_sheet in wb.sheetnames:
        del wb[neff_sheet]
    neff_ws = wb.create_sheet(neff_sheet)
    neff_headers = list(neff_rows[0].keys())
    neff_ws.append(neff_headers)
    for row in neff_rows:
        neff_ws.append([row[h] for h in neff_headers])

    index_ws = wb["00_Index"]
    existing_titles = {index_ws.cell(r, 1).value for r in range(2, index_ws.max_row + 1)}
    if neff_sheet not in existing_titles:
        index_ws.append([
            neff_sheet,
            "Factor-specific Neff estimates used to replace the original shared 599249 label in standard factor-GWAS text files.",
            None,
            "Neff was estimated as mean(1/(2*MAF*(1-MAF)*SE^2)) across SNPs with 0.1<=MAF<=0.4; this sheet records both the original label and the updated factor-specific values.",
        ])

    wb.save(TARGET_WORKBOOK)
    print(TARGET_WORKBOOK)


if __name__ == "__main__":
    main()
