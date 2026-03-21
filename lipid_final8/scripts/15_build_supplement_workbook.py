from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\metabolic\GWAS\genomicgem_main_zgt4_nonproportion")
OUT_DIR = ROOT / "supplement_lipid_final8"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_XLSX = OUT_DIR / "lipid_final8_supplement_workbook.xlsx"


def read_tsv(path):
    return pd.read_csv(path, sep="\t")


def read_csv_index(path):
    return pd.read_csv(path, index_col=0)


main_manifest = read_tsv(ROOT / "step1_ldsc_results" / "trait_manifest.tsv")
main_summary = read_tsv(ROOT / "step1_ldsc_results" / "Main_Zgt4_nonproportion_ldsc_summary.tsv")
main_qc = read_tsv(ROOT / "step1_ldsc_results" / "efa_selection" / "trait_qc_and_selection.tsv")
main_s = read_csv_index(ROOT / "step1_ldsc_results" / "Main_Zgt4_nonproportion_S_matrix.csv")
main_rg = read_csv_index(ROOT / "step1_ldsc_results" / "Main_Zgt4_nonproportion_rg_matrix.csv")

lipid94 = read_tsv(ROOT / "lipid_module_from_full_manifest" / "lipid_module_candidates.tsv")
compact20_review = read_tsv(ROOT / "lipid_module_from_full_manifest" / "compact_panel_review" / "lipid_module_compact_review.tsv")
ultra10_review = read_tsv(ROOT / "lipid_module_from_full_manifest" / "compact_panel_review" / "ultra_pure_3factor_review" / "lipid_module_ultrapure3_review.tsv")
final8_review = read_tsv(ROOT / "lipid_module_from_full_manifest" / "compact_panel_review" / "ultra_pure_3factor_review" / "final8_review" / "lipid_module_ultrapure3_final8_review.tsv")
final8_manifest = read_tsv(ROOT / "final_model_lipid_module_final8" / "final8_trait_manifest.tsv")

compact20_summary = read_tsv(ROOT / "step7_ldsc_lipid_module_compact20" / "lipid_module_compact20_ldsc_summary.tsv")
compact20_s = read_csv_index(ROOT / "step7_ldsc_lipid_module_compact20" / "lipid_module_compact20_S_matrix.csv")
compact20_rg = read_csv_index(ROOT / "step7_ldsc_lipid_module_compact20" / "lipid_module_compact20_rg_matrix.csv")
compact20_efa_fit = read_tsv(ROOT / "step8_efa_esem_lipid_module_compact20" / "EFA_minres_fit_summary.tsv")
compact20_esem_fit = read_tsv(ROOT / "step8_efa_esem_lipid_module_compact20" / "lipid_module_compact20_esem_target_summary.tsv")

final8_summary = read_tsv(ROOT / "step12_ldsc_lipid_module_final8" / "lipid_module_final8_ldsc_summary.tsv")
final8_s = read_csv_index(ROOT / "step12_ldsc_lipid_module_final8" / "lipid_module_final8_S_matrix.csv")
final8_rg = read_csv_index(ROOT / "step12_ldsc_lipid_module_final8" / "lipid_module_final8_rg_matrix.csv")

esem_fit = read_tsv(ROOT / "step13_efa_esem_lipid_module_final8" / "lipid_module_final8_esem_summary.tsv")
esem_loadings = read_tsv(ROOT / "step13_efa_esem_lipid_module_final8" / "ALL_3factor_loadings.tsv")
ultra10_efa_fit = read_tsv(ROOT / "step11_efa_esem_lipid_module_ultrapure3_10" / "EFA_minres_fit_summary.tsv")
ultra10_esem_fit = read_tsv(ROOT / "step11_efa_esem_lipid_module_ultrapure3_10" / "lipid_module_ultrapure3_10_esem_summary.tsv")
final8_efa_fit = read_tsv(ROOT / "step13_efa_esem_lipid_module_final8" / "EFA_minres_fit_summary.tsv")

fgwas_manifest = read_tsv(ROOT / "step14_native_wsl_usergwas_final8_results" / "merged_lipid_final8" / "standard_txt" / "lipid_final8_standard_txt_metadata.tsv")
factor_ldsc_uni = read_tsv(ROOT / "supplement_lipid_final8" / "Supplementary_Table_S5_lipid_final8_univariate_ldsc_detailed.tsv")
factor_ldsc_bi = read_tsv(ROOT / "supplement_lipid_final8" / "Supplementary_Table_S6_lipid_final8_bivariate_ldsc_detailed.tsv")
factor_ldsc_rg = pd.read_csv(ROOT / "step15_ldsc_validation_lipid_final8" / "lipid_final8_factor_rg_matrix.tsv", sep="\t")
factor_ldsc_rg.index = ["lipid8_F1", "lipid8_F2", "lipid8_F3"]


# Build lineage from 112 -> 94 -> 20 -> 10 -> 8
lineage = main_qc.merge(main_manifest, on=["study_accession", "trait_code", "biomarker_name", "group"], how="left")
lipid_set = set(lipid94["trait_code"])
compact_cols = [
    "trait_code", "selection_status", "selection_reason", "proposed_module",
    "marker_role", "high_rg_links_gt_0p95"
]
ultra_cols = [
    "trait_code", "selection_status", "selection_reason", "ultra_pure_factor",
    "same_primary", "mean_top_abs", "mean_gap"
]
final8_cols = [
    "trait_code", "selection_status_final8", "selection_reason_final8"
]

lineage["stage_94_status"] = lineage["trait_code"].map(lambda x: "kept" if x in lipid_set else "dropped")
lineage["stage_94_reason"] = lineage["group"].map(
    lambda g: "Retained in lipid-only candidate universe." if g in set(lipid94["group"]) else "Dropped because trait is outside the lipid-only module universe."
)

lineage = lineage.merge(
    compact20_review[compact_cols].rename(columns={
        "selection_status": "stage_20_status",
        "selection_reason": "stage_20_reason",
        "proposed_module": "stage_20_module",
        "marker_role": "stage_20_marker_role",
        "high_rg_links_gt_0p95": "stage_20_high_rg_links_gt_0p95",
    }),
    on="trait_code", how="left"
)
lineage = lineage.merge(
    ultra10_review[ultra_cols].rename(columns={
        "selection_status": "stage_10_status",
        "selection_reason": "stage_10_reason",
        "ultra_pure_factor": "stage_10_factor",
        "same_primary": "stage_10_same_primary",
        "mean_top_abs": "stage_10_mean_top_abs",
        "mean_gap": "stage_10_mean_gap",
    }),
    on="trait_code", how="left"
)
lineage = lineage.merge(final8_review[final8_cols], on="trait_code", how="left")
lineage = lineage.rename(columns={
    "selection_status_final8": "stage_8_status",
    "selection_reason_final8": "stage_8_reason",
})

for col in ["stage_20_status", "stage_10_status", "stage_8_status"]:
    lineage[col] = lineage[col].fillna("not_entered")

lineage["final_factor_name"] = lineage["trait_code"].map(
    dict(zip(final8_manifest["trait_code"], final8_manifest["final_factor_name"]))
)

criteria = pd.DataFrame([
    {
        "stage": "Stage 1",
        "input_n": "all Main_Zgt4_nonproportion traits",
        "output_n": 112,
        "selection_standard": "Retain traits with single-trait h2 Z > 4 in the original QC table.",
        "dropped_n": "not shown here",
        "notes": "This is the successful starting universe used for downstream multivariate LDSC."
    },
    {
        "stage": "Stage 2",
        "input_n": 112,
        "output_n": 94,
        "selection_standard": "Restrict to lipid-related groups only for module-specific modeling.",
        "dropped_n": 18,
        "notes": "Dropped non-lipid groups such as amino acids and other non-lipid metabolites."
    },
    {
        "stage": "Stage 3",
        "input_n": 94,
        "output_n": 20,
        "selection_standard": "Compact balanced panel preserving HDL, VLDL/TG, structural lipids, fatty acids, and bridge traits; prefer cleaner representatives under high rg redundancy.",
        "dropped_n": 74,
        "notes": "Retained panel had no pair with |rg| > 0.95."
    },
    {
        "stage": "Stage 4",
        "input_n": 20,
        "output_n": 10,
        "selection_standard": "Ultra-pure markers with stable primary factor in ALL and ODD, high absolute loading, large loading gap, and biological interpretability.",
        "dropped_n": 10,
        "notes": "This produced the successful 10-trait ultra-pure 3-factor panel."
    },
    {
        "stage": "Stage 5",
        "input_n": 10,
        "output_n": 8,
        "selection_standard": "Remove markers whose exclusion improved projected fit while preserving the 3-factor lipid structure.",
        "dropped_n": 2,
        "notes": "HDL_size and Total_CE were removed, yielding the final 8-trait lipid_final8 model."
    },
    {
        "stage": "Stage 6",
        "input_n": 8,
        "output_n": "3 factors",
        "selection_standard": "Final downstream model: 3-factor target-rotation ESEM followed by native WSL userGWAS and factor LDSC validation.",
        "dropped_n": 0,
        "notes": "Failed alternative models are intentionally excluded from this workbook."
    },
])

index_df = pd.DataFrame([
    ("00_Index", "Workbook directory and sheet descriptions."),
    ("01_StageCriteria", "Successful lipid_final8 selection stages from 112 traits to final 8 traits."),
    ("02_Main112_QC", "Initial 112-trait QC table (Z>4 universe) with h2/intercept and EFA-selection flag from the early QC file."),
    ("03_Main112_S", "112-trait multivariate LDSC S matrix."),
    ("04_Main112_rg", "112-trait multivariate LDSC genetic correlation matrix."),
    ("05_Lineage_112to8", "Trait-by-trait lineage across the 94, 20, 10, and final 8 selection stages."),
    ("06_Lipid94_Candidates", "Lipid-only candidate universe (94 traits)."),
    ("07_Compact20_Review", "20-trait compact lipid review with keep/drop decisions and reasons."),
    ("08_Compact20_LDSC", "20-trait compact panel multivariate LDSC summary."),
    ("09_Compact20_S", "20-trait compact panel S matrix."),
    ("10_Compact20_rg", "20-trait compact panel rg matrix."),
    ("11_UltraPure10_Review", "10-trait ultra-pure 3-factor review with keep/drop decisions and loading stability metrics."),
    ("12_Final8_Review", "Final 8-trait review showing which 10-trait markers were retained or removed."),
    ("13_Final8_Manifest", "Final 8-trait manifest and factor assignment."),
    ("14_Final8_LDSC", "Final 8-trait multivariate LDSC summary."),
    ("15_Final8_S", "Final 8-trait S matrix."),
    ("16_Final8_rg", "Final 8-trait rg matrix."),
    ("17_Final8_ESEM_Fit", "Final 8-trait 3-factor ESEM fit summary."),
    ("18_Final8_Loadings", "ALL dataset 3-factor loadings for the final 8-trait model."),
    ("19_Compact20_EFA_Fit", "Compact 20-trait EFA fit summary across factor numbers."),
    ("20_Compact20_ESEM_Fit", "Compact 20-trait ESEM model comparison summary."),
    ("21_Ultra10_EFA_Fit", "Ultra-pure 10-trait EFA fit summary."),
    ("22_Ultra10_ESEM_Fit", "Ultra-pure 10-trait ESEM model summary."),
    ("23_Final8_EFA_Fit", "Final 8-trait EFA fit summary."),
    ("24_FactorGWAS_Files", "Final merged userGWAS outputs and standard TXT exports."),
    ("25_FactorLDSC_Uni", "Final factor-level univariate LDSC summary with chi-square, GC, intercept, ratio, h2, and h2 Z."),
    ("26_FactorLDSC_Bi", "Final factor-level bivariate LDSC summary with g_cov and rg."),
    ("27_FactorLDSC_rg", "Final factor-level rg matrix.")
], columns=["sheet_name", "description"])


with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    index_df.to_excel(writer, sheet_name="00_Index", index=False)
    criteria.to_excel(writer, sheet_name="01_StageCriteria", index=False)
    main_qc.to_excel(writer, sheet_name="02_Main112_QC", index=False)
    main_s.to_excel(writer, sheet_name="03_Main112_S")
    main_rg.to_excel(writer, sheet_name="04_Main112_rg")
    lineage.to_excel(writer, sheet_name="05_Lineage_112to8", index=False)
    lipid94.to_excel(writer, sheet_name="06_Lipid94_Candidates", index=False)
    compact20_review.to_excel(writer, sheet_name="07_Compact20_Review", index=False)
    compact20_summary.to_excel(writer, sheet_name="08_Compact20_LDSC", index=False)
    compact20_s.to_excel(writer, sheet_name="09_Compact20_S")
    compact20_rg.to_excel(writer, sheet_name="10_Compact20_rg")
    ultra10_review.to_excel(writer, sheet_name="11_UltraPure10_Review", index=False)
    final8_review.to_excel(writer, sheet_name="12_Final8_Review", index=False)
    final8_manifest.to_excel(writer, sheet_name="13_Final8_Manifest", index=False)
    final8_summary.to_excel(writer, sheet_name="14_Final8_LDSC", index=False)
    final8_s.to_excel(writer, sheet_name="15_Final8_S")
    final8_rg.to_excel(writer, sheet_name="16_Final8_rg")
    esem_fit.to_excel(writer, sheet_name="17_Final8_ESEM_Fit", index=False)
    esem_loadings.to_excel(writer, sheet_name="18_Final8_Loadings", index=False)
    compact20_efa_fit.to_excel(writer, sheet_name="19_Compact20_EFA_Fit", index=False)
    compact20_esem_fit.to_excel(writer, sheet_name="20_Compact20_ESEM_Fit", index=False)
    ultra10_efa_fit.to_excel(writer, sheet_name="21_Ultra10_EFA_Fit", index=False)
    ultra10_esem_fit.to_excel(writer, sheet_name="22_Ultra10_ESEM_Fit", index=False)
    final8_efa_fit.to_excel(writer, sheet_name="23_Final8_EFA_Fit", index=False)
    fgwas_manifest.to_excel(writer, sheet_name="24_FactorGWAS_Files", index=False)
    factor_ldsc_uni.to_excel(writer, sheet_name="25_FactorLDSC_Uni", index=False)
    factor_ldsc_bi.to_excel(writer, sheet_name="26_FactorLDSC_Bi", index=False)
    factor_ldsc_rg.to_excel(writer, sheet_name="27_FactorLDSC_rg")


wb = load_workbook(OUT_XLSX)
header_fill = PatternFill("solid", fgColor="D9EAF7")
header_font = Font(bold=True)
top_align = Alignment(vertical="top", wrap_text=True)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = top_align
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells[:200]:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

if "00_Index" in wb.sheetnames:
    ws = wb["00_Index"]
    ws["A1"] = "sheet_name"
    ws["B1"] = "description"
    ws["D1"] = "notes"
    ws["D2"] = "Workbook includes only the successful lipid_final8 route. Failed model branches are intentionally excluded."
    ws["D3"] = "The 112-trait V matrix is not included because it exceeds practical Excel size/usefulness; S and rg matrices are included instead."
    ws["D2"].alignment = top_align
    ws["D3"].alignment = top_align
    ws.column_dimensions["D"].width = 80

wb.save(OUT_XLSX)
print(f"Wrote workbook: {OUT_XLSX}")
